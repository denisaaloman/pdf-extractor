from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

META_HEADERS = ["Fisier", "Tabel", "Categorie", "Rand"]

# Caractere interzise de Excel intr-un nume de sheet
_INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')


def _safe_sheet_title(name: str, used: set[str]) -> str:
    base = _INVALID_SHEET_CHARS.sub("", name or "Tabel").strip() or "Tabel"
    base = base[:28]
    title = base
    i = 2
    while title in used:
        suffix = f"_{i}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        i += 1
    used.add(title)
    return title


def _collect_unique_columns(results: list[dict[str, Any]]) -> list[str]:
    """Coloane unice, in ordinea primei aparitii, peste toate tabelele/documentele."""
    seen: dict[str, None] = {}
    for doc in results:
        for table in doc.get("tables", []):
            for col in table.get("columns") or []:
                seen.setdefault(col, None)
    return list(seen.keys())


def _iter_wide_rows(results: list[dict[str, Any]], unique_columns: list[str]):
    """
    Randuri in format wide: coloanele de identificare (Fisier/Tabel/Categorie/Rand)
    + o coloana pentru FIECARE coloana unica intalnita in oricare tabel.
    Pe un rand provenit dintr-un tabel care nu are o anumita coloana, celula
    ramane goala - asa se vede clar din ce tabel/pagina vine fiecare valoare,
    fara sa se amestece coloane intre tabele diferite.
    """
    col_index = {col: i for i, col in enumerate(unique_columns)}
    for doc in results:
        filename = doc.get("filename", "")
        for table in doc.get("tables", []):
            title = table.get("title", "")
            category = table.get("category", "")
            columns = table.get("columns") or []
            for row_i, row in enumerate(table.get("rows", []), start=1):
                values = [""] * len(unique_columns)
                for col_name, value in zip(columns, row):
                    values[col_index[col_name]] = value
                yield [filename, title, category, row_i] + values


def _autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0 for c in col_cells), default=8)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 60)


def build_xlsx(results: list[dict[str, Any]]) -> bytes:
    """
    results: lista de dict-uri {filename, tables: [{title, category, columns, rows}]}
    (formatul returnat de extraction.extract_tables_from_pdf)

    Un sheet PER TABEL/LISTA gasit(a). In fiecare sheet, header-ul e
    Fisier / Tabel / Categorie / Rand + coloanele originale ale tabelului,
    iar fiecare rand de date repeta Fisier/Tabel/Categorie/Rand ca si
    coloane efective (nu bloc de metadata separat).
    """
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()

    for doc in results:
        filename = doc.get("filename", "Document")
        for table in doc.get("tables", []):
            title = table.get("title") or "Tabel"
            category = table.get("category", "")
            columns = table.get("columns") or []
            rows = table.get("rows", [])

            sheet_name = _safe_sheet_title(title, used_titles)
            ws = wb.create_sheet(title=sheet_name)

            headers = META_HEADERS + columns
            for col_i, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_i, value=header).font = Font(bold=True)

            for row_i, row in enumerate(rows, start=1):
                line = [filename, title, category, row_i] + list(row)
                for col_i, value in enumerate(line, start=1):
                    ws.cell(row=row_i + 1, column=col_i, value=value)

            _autosize_columns(ws)
            ws.freeze_panes = "A2"

    if not wb.sheetnames:
        wb.create_sheet(title="Fara date")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(results: list[dict[str, Any]]) -> bytes:
    """CSV nu suporta mai multe sheet-uri - exportam formatul wide (toate tabelele intr-un singur tabel)."""
    unique_columns = _collect_unique_columns(results)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(META_HEADERS + unique_columns)
    for line in _iter_wide_rows(results, unique_columns):
        writer.writerow(line)

    return buf.getvalue().encode("utf-8-sig")  # BOM ca Excel sa deschida diacriticele corect